from flask import Flask, render_template, request, jsonify, send_file
import os, io, json, csv, time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import gspread
from google.oauth2.service_account import Credentials

app = Flask(__name__)

# ── Google Sheets Setup ───────────────────────────────────────────────────────
SHEET_NAME = os.environ.get('GOOGLE_SHEET_NAME', 'OrderTrack_DB')
creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
SHEET = None

if creds_json:
    try:
        creds_dict = json.loads(creds_json)
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(creds)
        SHEET = client.open(SHEET_NAME)
    except Exception as e:
        print(f"Failed to connect to Google Sheets: {e}")

# ── Cache ─────────────────────────────────────────────────────────────────────
_cache = {}
CACHE_TTL = 30

def cache_get(key):
    if key in _cache:
        data, ts = _cache[key]
        if time.time() - ts < CACHE_TTL: return data
    return None

def cache_set(key, data): _cache[key] = (data, time.time())
def cache_clear(key): _cache.pop(key, None)
def cache_clear_all(): _cache.clear()

# ── Helpers ───────────────────────────────────────────────────────────────────
def get_next_id(ws):
    try:
        col = ws.col_values(1)
        ids = [int(x) for x in col[1:] if str(x).isdigit()]
        return max(ids) + 1 if ids else 1
    except Exception:
        return 1

def safe_float(val, default=0.0):
    try:
        return float(val) if val not in (None, '', 'None') else default
    except (ValueError, TypeError):
        return default

def _safe_int(val, default=-1):
    """Safely cast a value to int — handles strings, floats, and None from gspread."""
    try:
        return int(val) if val not in (None, '', 'None') else default
    except (ValueError, TypeError):
        return default

# ── Page Routes ───────────────────────────────────────────────────────────────
@app.route('/')
def main_orders(): return render_template('main_orders.html')
@app.route('/secondary')
def secondary_orders(): return render_template('secondary_orders.html')
@app.route('/offline')
def offline_orders(): return render_template('offline_orders.html')
@app.route('/jiomart')
def jiomart_orders(): return render_template('jiomart_orders.html')
@app.route('/voucher-tracker')
def voucher_tracker(): return render_template('voucher_tracker.html')
@app.route('/exchange')
def exchange_orders(): return render_template('exchange_orders.html')
@app.route('/inventory')
def inventory(): return render_template('inventory.html')
@app.route('/dashboard')
def dashboard(): return render_template('dashboard.html')
@app.route('/settings')
def settings(): return render_template('settings.html')

# ── Settings APIs ─────────────────────────────────────────────────────────────
@app.route('/api/cards', methods=['GET', 'POST'])
def manage_cards():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('cards')
    if request.method == 'GET':
        cached = cache_get('cards')
        if cached: return jsonify(cached)
        try: cards = ws.get_all_records()
        except: cards = []
        for c in cards:
            if str(c.get('last_digits', '')).startswith("'"): c['last_digits'] = str(c['last_digits'])[1:]
        result = list(reversed(cards)); cache_set('cards', result); return jsonify(result)
    data = request.json; new_id = get_next_id(ws)
    last_digits = str(data.get('last_digits', ''))
    ws.append_row([new_id, data.get('card_type', ''), f"'{last_digits}" if last_digits else ""])
    cache_clear('cards'); cache_clear('card_lookup')
    return jsonify({'success': True})

@app.route('/api/cards/<int:card_id>', methods=['DELETE'])
def delete_card(card_id):
    ws = SHEET.worksheet('cards')
    try:
        cell = ws.find(str(card_id), in_column=1); ws.delete_rows(cell.row)
    except: pass
    cache_clear('cards'); cache_clear('card_lookup')
    return jsonify({'success': True})

@app.route('/api/card-lookup')
def card_lookup():
    digits = request.args.get('digits', '').strip()
    if digits.startswith("'"): digits = digits[1:]
    if not SHEET: return jsonify({'found': False})
    ws = SHEET.worksheet('cards')
    try: records = ws.get_all_records()
    except: records = []
    for row in records:
        db = str(row.get('last_digits', ''))
        if db.startswith("'"): db = db[1:]
        if db == digits: return jsonify({'card_type': row.get('card_type'), 'found': True})
    return jsonify({'found': False})

def handle_master_table(table_name, req, field_name='name'):
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet(table_name)
    if req.method == 'GET':
        cached = cache_get(f'master_{table_name}')
        if cached: return jsonify(cached)
        try: result = ws.get_all_records()
        except: result = []
        cache_set(f'master_{table_name}', result); return jsonify(result)
    data = req.json; new_id = get_next_id(ws)
    ws.append_row([new_id, data.get(field_name, '')])
    cache_clear(f'master_{table_name}')
    if table_name == 'platforms': cache_clear('platform_names')
    if table_name == 'models':
        for k in list(_cache.keys()):
            if k.startswith('variants'): cache_clear(k)
    return jsonify({'success': True})

def delete_master_table(table_name, item_id):
    ws = SHEET.worksheet(table_name)
    try:
        cell = ws.find(str(item_id), in_column=1); ws.delete_rows(cell.row)
    except: pass
    cache_clear(f'master_{table_name}')
    return jsonify({'success': True})

@app.route('/api/platforms', methods=['GET', 'POST'])
def api_platforms(): return handle_master_table('platforms', request, 'platform_name')
@app.route('/api/platforms/<int:id>', methods=['DELETE'])
def api_del_platforms(id): cache_clear('platform_names'); return delete_master_table('platforms', id)

@app.route('/api/models', methods=['GET', 'POST'])
def api_models(): return handle_master_table('models', request, 'model_name')
@app.route('/api/models/<int:id>', methods=['DELETE'])
def api_del_models(id): return delete_master_table('models', id)

@app.route('/api/sec-order-names', methods=['GET', 'POST'])
def api_sec_names(): return handle_master_table('sec_order_names', request, 'name')
@app.route('/api/sec-order-names/<int:id>', methods=['DELETE'])
def api_del_sec_names(id): return delete_master_table('sec_order_names', id)

@app.route('/api/machines', methods=['GET', 'POST'])
def api_machines(): return handle_master_table('machines', request, 'name')
@app.route('/api/machines/<int:id>', methods=['DELETE'])
def api_del_machines(id): return delete_master_table('machines', id)

@app.route('/api/vendors', methods=['GET', 'POST'])
def api_vendors(): return handle_master_table('vendors', request, 'name')
@app.route('/api/vendors/<int:id>', methods=['DELETE'])
def api_del_vendors(id): return delete_master_table('vendors', id)

@app.route('/api/brands', methods=['GET', 'POST'])
def api_brands(): return handle_master_table('brands', request, 'name')
@app.route('/api/brands/<int:id>', methods=['DELETE'])
def api_del_brands(id): return delete_master_table('brands', id)

# ── Vouchers ──────────────────────────────────────────────────────────────────
@app.route('/api/vouchers', methods=['GET', 'POST'])
def api_vouchers():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('vouchers')
    if request.method == 'GET':
        cached = cache_get('master_vouchers')
        if cached: return jsonify(cached)
        try: result = ws.get_all_records()
        except: result = []
        cache_set('master_vouchers', result); return jsonify(result)
    data = request.json; new_id = get_next_id(ws)
    ws.append_row([new_id, data.get('name', ''), safe_float(data.get('value', 0))])
    cache_clear('master_vouchers')
    return jsonify({'success': True})

@app.route('/api/vouchers/<int:id>', methods=['DELETE'])
def api_del_voucher(id): cache_clear('master_vouchers'); return delete_master_table('vouchers', id)

@app.route('/api/platform-names')
def platform_names():
    if not SHEET: return jsonify([])
    cached = cache_get('platform_names')
    if cached: return jsonify(cached)
    try: records = SHEET.worksheet('platforms').get_all_records()
    except: records = []
    result = sorted(list(set([r['platform_name'] for r in records if r.get('platform_name')])))
    cache_set('platform_names', result); return jsonify(result)

@app.route('/api/variants', methods=['GET', 'POST'])
def api_variants():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('variants')
    if request.method == 'GET':
        model_name = request.args.get('model')
        cache_key = f'variants_{model_name}' if model_name else 'variants_all'
        cached = cache_get(cache_key)
        if cached: return jsonify(cached)
        try: variants = ws.get_all_records()
        except: variants = []
        if model_name:
            try: models = SHEET.worksheet('models').get_all_records()
            except: models = []
            m_id = next((m['id'] for m in models if m['model_name'] == model_name), None)
            try: m_id_int = int(m_id) if m_id is not None else None
            except: m_id_int = None
            result = [v for v in variants if m_id_int is not None and _safe_int(v.get('model_id')) == m_id_int] if m_id_int else []
        else:
            result = variants
        cache_set(cache_key, result); return jsonify(result)
    data = request.json; new_id = get_next_id(ws)
    try: v_model_id_int = int(data.get('model_id', 0))
    except: v_model_id_int = 0
    ws.append_row([new_id, v_model_id_int, data.get('variant_name', ''), data.get('costing', ''), data.get('selling_price', '')])
    for k in list(_cache.keys()):
        if k.startswith('variants'): cache_clear(k)
    return jsonify({'success': True})

@app.route('/api/variants/<int:var_id>', methods=['DELETE', 'PUT'])
def del_variant(var_id):
    if request.method == 'DELETE':
        for k in list(_cache.keys()):
            if k.startswith('variants'): cache_clear(k)
        return delete_master_table('variants', var_id)
    if not SHEET: return jsonify({'success': False})
    data = request.json; ws = SHEET.worksheet('variants')
    try:
        cell = ws.find(str(var_id), in_column=1); row = ws.row_values(cell.row)
        new_cost = data.get('costing', row[3] if len(row) > 3 else '')
        new_sell = data.get('selling_price', row[4] if len(row) > 4 else '')
        ws.update(f'D{cell.row}:E{cell.row}', [[new_cost, new_sell]])
        for k in list(_cache.keys()):
            if k.startswith('variants'): cache_clear(k)
        return jsonify({'success': True, 'costing': new_cost, 'selling_price': new_sell})
    except Exception as e:
        print("Variant PUT error:", e); return jsonify({'success': False})

@app.route('/api/variants/<int:var_id>/sync-sell-price', methods=['POST'])
def sync_variant_sell_price(var_id):
    if not SHEET: return jsonify({'success': False})
    data = request.json
    variant_name = data.get('variant_name', ''); new_sell = data.get('selling_price', '')
    match_costing = data.get('costing', '')
    if not variant_name or new_sell == '': return jsonify({'success': False, 'error': 'missing fields'})
    try: new_sell_f = float(new_sell)
    except: return jsonify({'success': False, 'error': 'invalid selling_price'})
    try: match_cost_f = float(match_costing) if match_costing != '' else None
    except: match_cost_f = None
    updated = 0; errors = []
    # secondary_orders now has same column positions as main_orders for variant/costing/sell/profit
    for sheet_name, variant_col, sell_col, profit_col, cost_col in [
        ('main_orders',      8, 10, 11, 9),
        ('secondary_orders', 8, 10, 11, 9),
    ]:
        try:
            ws = SHEET.worksheet(sheet_name); all_rows = ws.get_all_values()
            if not all_rows: continue
            for row_idx, row in enumerate(all_rows[1:], start=2):
                cell_variant = row[variant_col - 1] if len(row) >= variant_col else ''
                if cell_variant.strip() != variant_name.strip(): continue
                try: row_cost_f = float(row[cost_col - 1]) if len(row) >= cost_col and row[cost_col - 1] else 0.0
                except: row_cost_f = 0.0
                if match_cost_f is not None and round(row_cost_f, 2) != round(match_cost_f, 2): continue
                ws.update_cell(row_idx, sell_col, new_sell_f)
                ws.update_cell(row_idx, profit_col, round(new_sell_f - row_cost_f, 2))
                updated += 1
        except Exception as e: errors.append(f"{sheet_name}: {e}")
    cache_clear('main_orders'); cache_clear('secondary_orders')
    return jsonify({'success': True, 'updated': updated, 'errors': errors})


# ── Main Orders ───────────────────────────────────────────────────────────────
# id(1) card_type(2) last_digits(3) platform(4) account(5) order_name(6)
# model(7) variant(8) costing(9) selling_price(10) profit(11) delivery_date(12)
# voucher_name(13) voucher_value(14) card_value(15) sale_month(16) created_at(17)

@app.route('/api/main-orders', methods=['GET', 'POST'])
def api_main_orders():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('main_orders')
    if request.method == 'GET':
        cached = cache_get('main_orders')
        if cached: return jsonify(cached)
        try: records = ws.get_all_records()
        except: records = []
        for o in records:
            if str(o.get('last_digits', '')).startswith("'"): o['last_digits'] = str(o['last_digits'])[1:]
        result = list(reversed(records)); cache_set('main_orders', result); return jsonify(result)
    try:
        data = request.json; next_id = get_next_id(ws)
        costing = safe_float(data.get('costing')); selling = safe_float(data.get('selling_price'))
        vv = safe_float(data.get('voucher_value')); cv = costing - vv
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ld = str(data.get('last_digits', '')); sd = f"'{ld}" if ld else ""
        ws.append_row([next_id, data.get('card_type',''), sd, data.get('platform',''),
            data.get('account',''), data.get('order_name',''), data.get('model',''),
            data.get('variant',''), costing, selling, selling-costing, data.get('delivery_date',''),
            data.get('voucher_name',''), vv if vv else '', cv if vv else '',
            data.get('sale_month',''), now])
        cache_clear('main_orders')
        return jsonify({'success':True,'id':next_id,'card_type':data.get('card_type',''),'last_digits':ld,
            'platform':data.get('platform',''),'account':data.get('account',''),
            'order_name':data.get('order_name',''),'model':data.get('model',''),
            'variant':data.get('variant',''),'costing':costing,'selling_price':selling,
            'profit':selling-costing,'delivery_date':data.get('delivery_date',''),
            'voucher_name':data.get('voucher_name',''),'voucher_value':vv if vv else '',
            'card_value':cv if vv else '','sale_month':data.get('sale_month',''),'created_at':now})
    except Exception as e:
        print(f"Main Orders POST Error: {e}"); return jsonify({'success': False}), 500

@app.route('/api/main-orders/bulk-delete', methods=['POST'])
def bulk_del_main():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids', []); ws = SHEET.worksheet('main_orders')
    try: records = ws.get_all_records()
    except: records = []
    rows_to_delete = [i+2 for i,r in enumerate(records) if r.get('id') in ids]
    for r_idx in sorted(rows_to_delete, reverse=True): ws.delete_row(r_idx)
    cache_clear('main_orders'); return jsonify({'success': True, 'deleted': len(rows_to_delete)})

@app.route('/api/main-orders/bulk-update-sale', methods=['POST'])
def bulk_sale_main():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids',[]); new_month = request.json.get('sale_month','')
    ws = SHEET.worksheet('main_orders')
    try:
        records = ws.get_all_records()
        for i,r in enumerate(records):
            if r.get('id') in ids: ws.update_cell(i+2, 16, new_month)
        cache_clear('main_orders'); return jsonify({'success': True})
    except Exception as e: print("Bulk Sale Error:", e); return jsonify({'success': False})

@app.route('/api/main-orders/bulk-set-sell', methods=['POST'])
def bulk_sell_main():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids',[]); new_sell = request.json.get('selling_price')
    if not ids or new_sell is None: return jsonify({'success': False})
    try: new_sell_f = float(new_sell)
    except: return jsonify({'success': False})
    ws = SHEET.worksheet('main_orders')
    try:
        records = ws.get_all_records(); updated = 0
        for i,r in enumerate(records):
            if r.get('id') in ids:
                cost = safe_float(r.get('costing'))
                ws.update_cell(i+2, 10, new_sell_f); ws.update_cell(i+2, 11, round(new_sell_f-cost,2)); updated+=1
        cache_clear('main_orders'); return jsonify({'success': True, 'updated': updated})
    except Exception as e: print("Bulk Sell Error:", e); return jsonify({'success': False})

@app.route('/api/main-orders/bulk-set-delivery', methods=['POST'])
def bulk_delivery_main():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids',[]); new_date = request.json.get('delivery_date','')
    if not ids or not new_date: return jsonify({'success': False})
    ws = SHEET.worksheet('main_orders')
    try:
        records = ws.get_all_records(); updated = 0
        for i,r in enumerate(records):
            if r.get('id') in ids: ws.update_cell(i+2, 12, new_date); updated+=1
        cache_clear('main_orders'); return jsonify({'success': True, 'updated': updated})
    except Exception as e: print("Bulk Delivery Error:", e); return jsonify({'success': False})

@app.route('/api/main-orders/export')
def export_main():
    if not SHEET: return "No sheet connected", 500
    fmt = request.args.get('format','csv'); sale_filter = request.args.get('sale','')
    try: records = SHEET.worksheet('main_orders').get_all_records()
    except: records = []
    for o in records:
        if str(o.get('last_digits','')).startswith("'"): o['last_digits']=str(o['last_digits'])[1:]
    if sale_filter and sale_filter!='ALL': records=[r for r in records if r.get('sale_month','')==sale_filter]
    headers=['id','card_type','last_digits','platform','account','order_name','model','variant',
             'costing','selling_price','profit','delivery_date','voucher_name','voucher_value',
             'card_value','sale_month','created_at']
    if fmt=='csv':
        out=io.StringIO(); w=csv.DictWriter(out,fieldnames=headers,extrasaction='ignore')
        w.writeheader(); w.writerows(records); out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode('utf-8')),mimetype='text/csv',
            as_attachment=True,download_name=f'main_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.csv')
    wb=openpyxl.Workbook(); ws_xl=wb.active; ws_xl.title="Main Orders"
    hf=PatternFill("solid",fgColor="1A2D45"); hfont=Font(bold=True,color="5BB8F5")
    ws_xl.append(headers)
    for c in ws_xl[1]: c.fill=hf; c.font=hfont
    for r in records: ws_xl.append([r.get(h,'') for h in headers])
    for col in ws_xl.columns: ws_xl.column_dimensions[col[0].column_letter].width=min(max((len(str(c.value or '')) for c in col),default=10)+4,40)
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,download_name=f'main_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')

@app.route('/api/main-orders/<int:id>', methods=['DELETE', 'PUT'])
def modify_main(id):
    if request.method=='DELETE': cache_clear('main_orders'); return delete_master_table('main_orders',id)
    if not SHEET: return jsonify({'success': False})
    data=request.json; ws=SHEET.worksheet('main_orders')
    try:
        cell=ws.find(str(id),in_column=1)
        cost=safe_float(data.get('costing')); sell=safe_float(data.get('selling_price'))
        vv=safe_float(data.get('voucher_value')); cv=cost-vv if vv else ''
        ld=str(data.get('last_digits','')); sd=f"'{ld}" if ld else ""
        ws.update(f'B{cell.row}:P{cell.row}',[[data.get('card_type',''),sd,data.get('platform',''),
            data.get('account',''),data.get('order_name',''),data.get('model',''),data.get('variant',''),
            cost,sell,sell-cost,data.get('delivery_date',''),
            data.get('voucher_name',''),vv if vv else '',cv,data.get('sale_month','')]])
        cache_clear('main_orders')
        return jsonify({'success':True,'id':id,'card_type':data.get('card_type',''),'last_digits':ld,
            'platform':data.get('platform',''),'account':data.get('account',''),
            'order_name':data.get('order_name',''),'model':data.get('model',''),
            'variant':data.get('variant',''),'costing':cost,'selling_price':sell,'profit':sell-cost,
            'delivery_date':data.get('delivery_date',''),'voucher_name':data.get('voucher_name',''),
            'voucher_value':vv if vv else '','card_value':cv,'sale_month':data.get('sale_month','')})
    except Exception as e: print("Edit Error:",e); return jsonify({'success': False})


# ── Secondary Orders ──────────────────────────────────────────────────────────
# id(1) card_type(2) last_digits(3) platform(4) ordered_by(5) order_name(6)
# model(7) variant(8) costing(9) selling_price(10) profit(11) delivery_date(12)
# voucher_name(13) voucher_value(14) card_value(15) sale_month(16) created_at(17)

@app.route('/api/secondary-orders', methods=['GET', 'POST'])
def api_secondary_orders():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('secondary_orders')
    if request.method=='GET':
        cached=cache_get('secondary_orders')
        if cached: return jsonify(cached)
        try: records=ws.get_all_records()
        except: records=[]
        for o in records:
            if str(o.get('last_digits','')).startswith("'"): o['last_digits']=str(o['last_digits'])[1:]
        result=list(reversed(records)); cache_set('secondary_orders',result); return jsonify(result)
    try:
        data=request.json; next_id=get_next_id(ws)
        costing=safe_float(data.get('costing')); selling=safe_float(data.get('selling_price'))
        vv=safe_float(data.get('voucher_value')); cv=costing-vv
        now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ld=str(data.get('last_digits','')); sd=f"'{ld}" if ld else ""
        ws.append_row([next_id,data.get('card_type',''),sd,data.get('platform',''),
            data.get('ordered_by',''),data.get('order_name',''),
            data.get('model',''),data.get('variant',''),
            costing,selling,selling-costing,data.get('delivery_date',''),
            data.get('voucher_name',''),vv if vv else '',cv if vv else '',
            data.get('sale_month',''),now])
        cache_clear('secondary_orders')
        return jsonify({'success':True,'id':next_id,'card_type':data.get('card_type',''),'last_digits':ld,
            'platform':data.get('platform',''),'ordered_by':data.get('ordered_by',''),
            'order_name':data.get('order_name',''),'model':data.get('model',''),
            'variant':data.get('variant',''),'costing':costing,'selling_price':selling,
            'profit':selling-costing,'delivery_date':data.get('delivery_date',''),
            'voucher_name':data.get('voucher_name',''),'voucher_value':vv if vv else '',
            'card_value':cv if vv else '','sale_month':data.get('sale_month',''),'created_at':now})
    except Exception as e: print(f"Secondary Orders POST Error: {e}"); return jsonify({'success': False}), 500

@app.route('/api/secondary-orders/bulk-delete', methods=['POST'])
def bulk_del_sec():
    if not SHEET: return jsonify({'success': False})
    ids=request.json.get('ids',[]); ws=SHEET.worksheet('secondary_orders')
    try: records=ws.get_all_records()
    except: records=[]
    rows_to_delete=[i+2 for i,r in enumerate(records) if r.get('id') in ids]
    for r_idx in sorted(rows_to_delete,reverse=True): ws.delete_row(r_idx)
    cache_clear('secondary_orders'); return jsonify({'success':True,'deleted':len(rows_to_delete)})

@app.route('/api/secondary-orders/bulk-update-sale', methods=['POST'])
def bulk_sale_sec():
    if not SHEET: return jsonify({'success': False})
    ids=request.json.get('ids',[]); new_month=request.json.get('sale_month','Current Sale')
    ws=SHEET.worksheet('secondary_orders')
    try:
        records=ws.get_all_records()
        for i,r in enumerate(records):
            if r.get('id') in ids: ws.update_cell(i+2,16,new_month)
        cache_clear('secondary_orders'); return jsonify({'success': True})
    except Exception as e: print("Bulk Sale Error:",e); return jsonify({'success': False})

@app.route('/api/secondary-orders/bulk-set-sell', methods=['POST'])
def bulk_sell_sec():
    if not SHEET: return jsonify({'success': False})
    ids=request.json.get('ids',[]); new_sell=request.json.get('selling_price')
    if not ids or new_sell is None: return jsonify({'success': False})
    try: new_sell_f=float(new_sell)
    except: return jsonify({'success': False})
    ws=SHEET.worksheet('secondary_orders')
    try:
        records=ws.get_all_records(); updated=0
        for i,r in enumerate(records):
            if r.get('id') in ids:
                cost=safe_float(r.get('costing'))
                ws.update_cell(i+2,10,new_sell_f); ws.update_cell(i+2,11,round(new_sell_f-cost,2)); updated+=1
        cache_clear('secondary_orders'); return jsonify({'success':True,'updated':updated})
    except Exception as e: print("Bulk Sell Sec Error:",e); return jsonify({'success': False})

@app.route('/api/secondary-orders/bulk-set-delivery', methods=['POST'])
def bulk_delivery_sec():
    if not SHEET: return jsonify({'success': False})
    ids=request.json.get('ids',[]); new_date=request.json.get('delivery_date','')
    if not ids or not new_date: return jsonify({'success': False})
    ws=SHEET.worksheet('secondary_orders')
    try:
        records=ws.get_all_records(); updated=0
        for i,r in enumerate(records):
            if r.get('id') in ids: ws.update_cell(i+2,12,new_date); updated+=1
        cache_clear('secondary_orders'); return jsonify({'success':True,'updated':updated})
    except Exception as e: print("Bulk Delivery Sec Error:",e); return jsonify({'success': False})

@app.route('/api/secondary-orders/export')
def export_secondary():
    if not SHEET: return "No sheet connected", 500
    fmt=request.args.get('format','csv'); sale_filter=request.args.get('sale','')
    try: records=SHEET.worksheet('secondary_orders').get_all_records()
    except: records=[]
    for o in records:
        if str(o.get('last_digits','')).startswith("'"): o['last_digits']=str(o['last_digits'])[1:]
    if sale_filter and sale_filter!='ALL': records=[r for r in records if r.get('sale_month','')==sale_filter]
    headers=['id','card_type','last_digits','platform','ordered_by','order_name','model','variant',
             'costing','selling_price','profit','delivery_date','voucher_name','voucher_value',
             'card_value','sale_month','created_at']
    if fmt=='csv':
        out=io.StringIO(); w=csv.DictWriter(out,fieldnames=headers,extrasaction='ignore')
        w.writeheader(); w.writerows(records); out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode('utf-8')),mimetype='text/csv',
            as_attachment=True,download_name=f'secondary_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.csv')
    wb=openpyxl.Workbook(); ws_xl=wb.active; ws_xl.title="Secondary Orders"
    hf=PatternFill("solid",fgColor="1A2D45"); hfont=Font(bold=True,color="5BB8F5")
    ws_xl.append(headers)
    for c in ws_xl[1]: c.fill=hf; c.font=hfont
    for r in records: ws_xl.append([r.get(h,'') for h in headers])
    for col in ws_xl.columns: ws_xl.column_dimensions[col[0].column_letter].width=min(max((len(str(c.value or '')) for c in col),default=10)+4,40)
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,download_name=f'secondary_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')

@app.route('/api/secondary-orders/<int:id>', methods=['DELETE', 'PUT'])
def modify_sec(id):
    if request.method=='DELETE': cache_clear('secondary_orders'); return delete_master_table('secondary_orders',id)
    if not SHEET: return jsonify({'success': False})
    data=request.json; ws=SHEET.worksheet('secondary_orders')
    try:
        cell=ws.find(str(id),in_column=1)
        cost=safe_float(data.get('costing')); sell=safe_float(data.get('selling_price'))
        vv=safe_float(data.get('voucher_value')); cv=cost-vv if vv else ''
        ld=str(data.get('last_digits','')); sd=f"'{ld}" if ld else ""
        # B(2) through P(16) = 15 values
        ws.update(f'B{cell.row}:P{cell.row}',[[data.get('card_type',''),sd,data.get('platform',''),
            data.get('ordered_by',''),data.get('order_name',''),
            data.get('model',''),data.get('variant',''),
            cost,sell,sell-cost,data.get('delivery_date',''),
            data.get('voucher_name',''),vv if vv else '',cv,data.get('sale_month','')]])
        cache_clear('secondary_orders')
        return jsonify({'success':True,'id':id,'card_type':data.get('card_type',''),'last_digits':ld,
            'platform':data.get('platform',''),'ordered_by':data.get('ordered_by',''),
            'order_name':data.get('order_name',''),'model':data.get('model',''),
            'variant':data.get('variant',''),'costing':cost,'selling_price':sell,'profit':sell-cost,
            'delivery_date':data.get('delivery_date',''),'voucher_name':data.get('voucher_name',''),
            'voucher_value':vv if vv else '','card_value':cv,'sale_month':data.get('sale_month','')})
    except Exception as e: print("Edit Error:",e); return jsonify({'success': False})


# ── Offline Orders ────────────────────────────────────────────────────────────
@app.route('/api/offline-orders', methods=['GET', 'POST'])
def api_offline_orders():
    if not SHEET: return jsonify([])
    ws=SHEET.worksheet('offline_orders')
    if request.method=='GET':
        cached=cache_get('offline_orders')
        if cached: return jsonify(cached)
        try: records=ws.get_all_records()
        except: records=[]
        for o in records:
            if str(o.get('last_digits','')).startswith("'"): o['last_digits']=str(o['last_digits'])[1:]
        result=list(reversed(records)); cache_set('offline_orders',result); return jsonify(result)
    try:
        data=request.json; next_id=get_next_id(ws)
        costing=safe_float(data.get('costing')); selling=safe_float(data.get('selling_price'))
        now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ld=str(data.get('last_digits','')); sd=f"'{ld}" if ld else ""
        ws.append_row([next_id,data.get('card_type',''),sd,data.get('machine',''),
            data.get('vendor',''),data.get('brand',''),data.get('sale_type',''),
            costing,selling,selling-costing,data.get('sale_month',''),now])
        cache_clear('offline_orders')
        return jsonify({'success':True,'id':next_id,'card_type':data.get('card_type',''),'last_digits':ld,
            'machine':data.get('machine',''),'vendor':data.get('vendor',''),'brand':data.get('brand',''),
            'sale_type':data.get('sale_type',''),'costing':costing,'selling_price':selling,
            'profit':selling-costing,'sale_month':data.get('sale_month',''),'created_at':now})
    except Exception as e: print(f"Offline POST Error: {e}"); return jsonify({'success': False}), 500

@app.route('/api/offline-orders/bulk-delete', methods=['POST'])
def bulk_del_offline():
    if not SHEET: return jsonify({'success': False})
    ids=request.json.get('ids',[]); ws=SHEET.worksheet('offline_orders')
    try: records=ws.get_all_records()
    except: records=[]
    rows_to_delete=[i+2 for i,r in enumerate(records) if r.get('id') in ids]
    for r_idx in sorted(rows_to_delete,reverse=True): ws.delete_row(r_idx)
    cache_clear('offline_orders'); return jsonify({'success':True,'deleted':len(rows_to_delete)})

@app.route('/api/offline-orders/bulk-set-costing', methods=['POST'])
def bulk_costing_offline():
    if not SHEET: return jsonify({'success': False})
    ids=request.json.get('ids',[]); new_cost=request.json.get('costing')
    if not ids or new_cost is None: return jsonify({'success': False})
    try: new_cost_f=float(new_cost)
    except: return jsonify({'success': False})
    ws=SHEET.worksheet('offline_orders')
    try:
        records=ws.get_all_records(); updated=0
        for i,r in enumerate(records):
            if r.get('id') in ids:
                sell=safe_float(r.get('selling_price'))
                ws.update_cell(i+2,8,new_cost_f); ws.update_cell(i+2,10,round(sell-new_cost_f,2)); updated+=1
        cache_clear('offline_orders'); return jsonify({'success':True,'updated':updated})
    except Exception as e: print("Bulk Costing Offline Error:",e); return jsonify({'success': False})

@app.route('/api/offline-orders/bulk-set-sell', methods=['POST'])
def bulk_sell_offline():
    if not SHEET: return jsonify({'success': False})
    ids=request.json.get('ids',[]); new_sell=request.json.get('selling_price')
    if not ids or new_sell is None: return jsonify({'success': False})
    try: new_sell_f=float(new_sell)
    except: return jsonify({'success': False})
    ws=SHEET.worksheet('offline_orders')
    try:
        records=ws.get_all_records(); updated=0
        for i,r in enumerate(records):
            if r.get('id') in ids:
                cost=safe_float(r.get('costing'))
                ws.update_cell(i+2,9,new_sell_f); ws.update_cell(i+2,10,round(new_sell_f-cost,2)); updated+=1
        cache_clear('offline_orders'); return jsonify({'success':True,'updated':updated})
    except Exception as e: print("Bulk Sell Offline Error:",e); return jsonify({'success': False})

@app.route('/api/offline-orders/export')
def export_offline():
    if not SHEET: return "No sheet connected", 500
    fmt=request.args.get('format','csv'); month_filter=request.args.get('month','')
    try: records=SHEET.worksheet('offline_orders').get_all_records()
    except: records=[]
    for o in records:
        if str(o.get('last_digits','')).startswith("'"): o['last_digits']=str(o['last_digits'])[1:]
    if month_filter: records=[r for r in records if r.get('sale_month','')==month_filter]
    headers=['id','card_type','last_digits','machine','vendor','brand','sale_type','costing','selling_price','profit','sale_month','created_at']
    if fmt=='csv':
        out=io.StringIO(); w=csv.DictWriter(out,fieldnames=headers,extrasaction='ignore')
        w.writeheader(); w.writerows(records); out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode('utf-8')),mimetype='text/csv',
            as_attachment=True,download_name=f'offline_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.csv')
    wb=openpyxl.Workbook(); ws_xl=wb.active; ws_xl.title="Offline Sales"
    hf=PatternFill("solid",fgColor="1A2D45"); hfont=Font(bold=True,color="2ECC8F")
    ws_xl.append(headers)
    for c in ws_xl[1]: c.fill=hf; c.font=hfont
    for r in records: ws_xl.append([r.get(h,'') for h in headers])
    for col in ws_xl.columns: ws_xl.column_dimensions[col[0].column_letter].width=min(max((len(str(c.value or '')) for c in col),default=10)+4,40)
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,download_name=f'offline_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')

@app.route('/api/offline-orders/<int:id>', methods=['DELETE', 'PUT'])
def modify_offline(id):
    if request.method=='DELETE': cache_clear('offline_orders'); return delete_master_table('offline_orders',id)
    if not SHEET: return jsonify({'success': False})
    data=request.json; ws=SHEET.worksheet('offline_orders')
    try:
        cell=ws.find(str(id),in_column=1)
        cost=safe_float(data.get('costing')); sell=safe_float(data.get('selling_price'))
        ld=str(data.get('last_digits','')); sd=f"'{ld}" if ld else ""
        ws.update(f'B{cell.row}:K{cell.row}',[[data.get('card_type',''),sd,data.get('machine',''),
            data.get('vendor',''),data.get('brand',''),data.get('sale_type',''),
            cost,sell,sell-cost,data.get('sale_month','')]])
        cache_clear('offline_orders')
        return jsonify({'success':True,'id':id,'card_type':data.get('card_type',''),'last_digits':ld,
            'machine':data.get('machine',''),'vendor':data.get('vendor',''),'brand':data.get('brand',''),
            'sale_type':data.get('sale_type',''),'costing':cost,'selling_price':sell,
            'profit':sell-cost,'sale_month':data.get('sale_month','')})
    except Exception as e: print("Edit Error:",e); return jsonify({'success': False})


# ── Schema & Setup ────────────────────────────────────────────────────────────
SHEET_SCHEMA = {
    'main_orders': ['id','card_type','last_digits','platform','account','order_name',
        'model','variant','costing','selling_price','profit','delivery_date',
        'voucher_name','voucher_value','card_value','sale_month','created_at'],
    'secondary_orders': ['id','card_type','last_digits','platform','ordered_by','order_name',
        'model','variant','costing','selling_price','profit','delivery_date',
        'voucher_name','voucher_value','card_value','sale_month','created_at'],
    'offline_orders': ['id','card_type','last_digits','machine','vendor','brand',
        'sale_type','costing','selling_price','profit','sale_month','created_at'],
    'cards':           ['id','card_type','last_digits'],
    'platforms':       ['id','platform_name','account_name'],
    'models':          ['id','model_name'],
    'variants':        ['id','model_id','variant_name','costing','selling_price'],
    'sec_order_names': ['id','name'],
    'vouchers':        ['id','name','value'],
    'machines':        ['id','name'],
    'vendors':         ['id','name'],
    'brands':           ['id','name'],
    'jiomart_orders':   ['id','card_type','last_digits','account','order_name','order_id',
                         'model','variant','costing','selling_price','profit',
                         'delivery_date','sale_month','created_at'],
    'jiomart_accounts': ['id','name'],
    'jiomart_models':   ['id','model_name'],
    'jiomart_variants': ['id','model_id','variant_name','costing','selling_price'],
    'voucher_tracker':  ['id','platform','voucher_code','voucher_pin','amount','discount_pct','profit','month','is_redeemed','created_at'],
    'voucher_commission': ['id','month','commission_amount','notes','created_at'],
    'exchange_orders':  ['id','platform','model','variant','costing',
                         'exchange_model','exchange_variant','exchange_value',
                         'service_fee','original_costing','last_digits','card_type',
                         'voucher_amount','created_at'],
}


# ── Jiomart Master Data APIs ─────────────────────────────────────────────────

@app.route('/api/jiomart-accounts', methods=['GET', 'POST'])
def api_jiomart_accounts(): return handle_master_table('jiomart_accounts', request, 'name')
@app.route('/api/jiomart-accounts/<int:id>', methods=['DELETE'])
def api_del_jiomart_accounts(id):
    cache_clear('master_jiomart_accounts')
    return delete_master_table('jiomart_accounts', id)

@app.route('/api/jiomart-models', methods=['GET', 'POST'])
def api_jiomart_models(): return handle_master_table('jiomart_models', request, 'model_name')
@app.route('/api/jiomart-models/<int:id>', methods=['DELETE'])
def api_del_jiomart_models(id):
    cache_clear('master_jiomart_models')
    return delete_master_table('jiomart_models', id)

@app.route('/api/jiomart-variants', methods=['GET', 'POST'])
def api_jiomart_variants():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('jiomart_variants')
    if request.method == 'GET':
        model_name = request.args.get('model')
        cache_key = f'jiomart_variants_{model_name}' if model_name else 'jiomart_variants_all'
        cached = cache_get(cache_key)
        if cached: return jsonify(cached)
        try: variants = ws.get_all_records()
        except: variants = []
        if model_name:
            try: models = SHEET.worksheet('jiomart_models').get_all_records()
            except: models = []
            m_id = next((m['id'] for m in models if m['model_name'] == model_name), None)
            # Cast both sides to int — gspread can return strings or ints inconsistently
            try: m_id_int = int(m_id) if m_id is not None else None
            except: m_id_int = None
            result = [v for v in variants if m_id_int is not None and _safe_int(v.get('model_id')) == m_id_int] if m_id_int else []
        else:
            result = variants
        cache_set(cache_key, result)
        return jsonify(result)
    # POST — always store model_id as integer to avoid string/int mismatch on read
    data = request.json; new_id = get_next_id(ws)
    try: model_id_int = int(data.get('model_id', 0))
    except: model_id_int = 0
    ws.append_row([new_id, model_id_int, data.get('variant_name',''),
                   data.get('costing',''), data.get('selling_price','')])
    for k in list(_cache.keys()):
        if k.startswith('jiomart_variants'): cache_clear(k)
    return jsonify({'success': True})

@app.route('/api/jiomart-variants/<int:var_id>', methods=['DELETE', 'PUT'])
def modify_jiomart_variant(var_id):
    if request.method == 'DELETE':
        for k in list(_cache.keys()):
            if k.startswith('jiomart_variants'): cache_clear(k)
        return delete_master_table('jiomart_variants', var_id)
    if not SHEET: return jsonify({'success': False})
    data = request.json; ws = SHEET.worksheet('jiomart_variants')
    try:
        cell = ws.find(str(var_id), in_column=1); row = ws.row_values(cell.row)
        new_cost = data.get('costing',       row[3] if len(row) > 3 else '')
        new_sell = data.get('selling_price', row[4] if len(row) > 4 else '')
        ws.update(f'D{cell.row}:E{cell.row}', [[new_cost, new_sell]])
        for k in list(_cache.keys()):
            if k.startswith('jiomart_variants'): cache_clear(k)
        return jsonify({'success': True, 'costing': new_cost, 'selling_price': new_sell})
    except Exception as e:
        print("Jiomart Variant PUT error:", e); return jsonify({'success': False})

@app.route('/api/jiomart-variants/<int:var_id>/sync-sell-price', methods=['POST'])
def sync_jiomart_variant_sell_price(var_id):
    if not SHEET: return jsonify({'success': False})
    data = request.json
    variant_name  = data.get('variant_name', '')
    new_sell      = data.get('selling_price', '')
    match_costing = data.get('costing', '')
    if not variant_name or new_sell == '':
        return jsonify({'success': False, 'error': 'variant_name and selling_price required'})
    try: new_sell_f = float(new_sell)
    except: return jsonify({'success': False, 'error': 'invalid selling_price'})
    try: match_cost_f = float(match_costing) if match_costing != '' else None
    except: match_cost_f = None
    updated = 0; errors = []
    # jiomart_orders: variant=H(8), costing=I(9), sell=J(10), profit=K(11)
    try:
        ws = SHEET.worksheet('jiomart_orders'); all_rows = ws.get_all_values()
        if all_rows:
            for row_idx, row in enumerate(all_rows[1:], start=2):
                cell_variant = row[7] if len(row) >= 8 else ''
                if cell_variant.strip() != variant_name.strip(): continue
                try: row_cost_f = float(row[8]) if len(row) >= 9 and row[8] else 0.0
                except: row_cost_f = 0.0
                if match_cost_f is not None and round(row_cost_f,2) != round(match_cost_f,2): continue
                ws.update_cell(row_idx, 10, new_sell_f)
                ws.update_cell(row_idx, 11, round(new_sell_f - row_cost_f, 2))
                updated += 1
    except Exception as e: errors.append(f"jiomart_orders: {e}")
    cache_clear('jiomart_orders')
    return jsonify({'success': True, 'updated': updated, 'errors': errors})


# ── Jiomart Orders API ────────────────────────────────────────────────────────
# id(1) card_type(2) last_digits(3) account(4) order_name(5) order_id(6)
# model(7) variant(8) costing(9) selling_price(10) profit(11)
# delivery_date(12) sale_month(13) created_at(14)

@app.route('/api/jiomart-orders', methods=['GET', 'POST'])
def api_jiomart_orders():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('jiomart_orders')

    if request.method == 'GET':
        cached = cache_get('jiomart_orders')
        if cached: return jsonify(cached)
        try: records = ws.get_all_records()
        except: records = []
        for o in records:
            if str(o.get('last_digits','')).startswith("'"): o['last_digits'] = str(o['last_digits'])[1:]
        result = list(reversed(records)); cache_set('jiomart_orders', result); return jsonify(result)

    try:
        data = request.json; next_id = get_next_id(ws)
        costing = safe_float(data.get('costing')); selling = safe_float(data.get('selling_price'))
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ld = str(data.get('last_digits', '')); sd = f"'{ld}" if ld else ""
        ws.append_row([
            next_id, data.get('card_type',''), sd, data.get('account',''),
            data.get('order_name',''), data.get('order_id',''),
            data.get('model',''), data.get('variant',''),
            costing, selling, selling - costing,
            data.get('delivery_date',''), data.get('sale_month',''), now
        ])
        cache_clear('jiomart_orders')
        return jsonify({
            'success': True, 'id': next_id,
            'card_type': data.get('card_type',''), 'last_digits': ld,
            'account': data.get('account',''),
            'order_name': data.get('order_name',''), 'order_id': data.get('order_id',''),
            'model': data.get('model',''), 'variant': data.get('variant',''),
            'costing': costing, 'selling_price': selling, 'profit': selling - costing,
            'delivery_date': data.get('delivery_date',''),
            'sale_month': data.get('sale_month',''), 'created_at': now
        })
    except Exception as e:
        print(f"Jiomart POST Error: {e}"); return jsonify({'success': False}), 500

@app.route('/api/jiomart-orders/bulk-delete', methods=['POST'])
def bulk_del_jiomart():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids', []); ws = SHEET.worksheet('jiomart_orders')
    try: records = ws.get_all_records()
    except: records = []
    rows_to_delete = [i+2 for i,r in enumerate(records) if r.get('id') in ids]
    for r_idx in sorted(rows_to_delete, reverse=True): ws.delete_row(r_idx)
    cache_clear('jiomart_orders'); return jsonify({'success': True, 'deleted': len(rows_to_delete)})

@app.route('/api/jiomart-orders/bulk-update-sale', methods=['POST'])
def bulk_sale_jiomart():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids', []); new_month = request.json.get('sale_month', 'Current Sale')
    ws = SHEET.worksheet('jiomart_orders')
    try:
        records = ws.get_all_records()
        for i, r in enumerate(records):
            if r.get('id') in ids: ws.update_cell(i+2, 13, new_month)  # col 13 = sale_month
        cache_clear('jiomart_orders'); return jsonify({'success': True})
    except Exception as e: print("Jiomart Bulk Sale Error:", e); return jsonify({'success': False})

@app.route('/api/jiomart-orders/bulk-set-sell', methods=['POST'])
def bulk_sell_jiomart():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids', []); new_sell = request.json.get('selling_price')
    if not ids or new_sell is None: return jsonify({'success': False})
    try: new_sell_f = float(new_sell)
    except: return jsonify({'success': False})
    ws = SHEET.worksheet('jiomart_orders')
    try:
        records = ws.get_all_records(); updated = 0
        for i, r in enumerate(records):
            if r.get('id') in ids:
                cost = safe_float(r.get('costing'))
                ws.update_cell(i+2, 10, new_sell_f)               # selling_price col 10
                ws.update_cell(i+2, 11, round(new_sell_f-cost,2)) # profit col 11
                updated += 1
        cache_clear('jiomart_orders'); return jsonify({'success': True, 'updated': updated})
    except Exception as e: print("Jiomart Bulk Sell Error:", e); return jsonify({'success': False})

@app.route('/api/jiomart-orders/bulk-set-delivery', methods=['POST'])
def bulk_delivery_jiomart():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids', []); new_date = request.json.get('delivery_date', '')
    if not ids or not new_date: return jsonify({'success': False})
    ws = SHEET.worksheet('jiomart_orders')
    try:
        records = ws.get_all_records(); updated = 0
        for i, r in enumerate(records):
            if r.get('id') in ids:
                ws.update_cell(i+2, 12, new_date)  # delivery_date col 12
                updated += 1
        cache_clear('jiomart_orders'); return jsonify({'success': True, 'updated': updated})
    except Exception as e: print("Jiomart Bulk Delivery Error:", e); return jsonify({'success': False})

@app.route('/api/jiomart-orders/export')
def export_jiomart():
    if not SHEET: return "No sheet connected", 500
    fmt = request.args.get('format', 'csv'); sale_filter = request.args.get('sale', '')
    try: records = SHEET.worksheet('jiomart_orders').get_all_records()
    except: records = []
    for o in records:
        if str(o.get('last_digits','')).startswith("'"): o['last_digits'] = str(o['last_digits'])[1:]
    if sale_filter and sale_filter != 'ALL':
        records = [r for r in records if r.get('sale_month','') == sale_filter]
    headers = ['id','card_type','last_digits','account','order_name','order_id',
               'model','variant','costing','selling_price','profit',
               'delivery_date','sale_month','created_at']
    if fmt == 'csv':
        out = io.StringIO(); w = csv.DictWriter(out, fieldnames=headers, extrasaction='ignore')
        w.writeheader(); w.writerows(records); out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode('utf-8')), mimetype='text/csv',
            as_attachment=True, download_name=f'jiomart_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.csv')
    wb = openpyxl.Workbook(); ws_xl = wb.active; ws_xl.title = "Jiomart Orders"
    hf = PatternFill("solid", fgColor="1A2D45"); hfont = Font(bold=True, color="F5A623")
    ws_xl.append(headers)
    for c in ws_xl[1]: c.fill = hf; c.font = hfont
    for r in records: ws_xl.append([r.get(h,'') for h in headers])
    for col in ws_xl.columns:
        ws_xl.column_dimensions[col[0].column_letter].width = min(max((len(str(c.value or '')) for c in col), default=10)+4, 40)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'jiomart_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')

@app.route('/api/jiomart-orders/<int:id>', methods=['DELETE', 'PUT'])
def modify_jiomart(id):
    if request.method == 'DELETE':
        cache_clear('jiomart_orders'); return delete_master_table('jiomart_orders', id)
    if not SHEET: return jsonify({'success': False})
    data = request.json; ws = SHEET.worksheet('jiomart_orders')
    try:
        cell = ws.find(str(id), in_column=1)
        cost = safe_float(data.get('costing')); sell = safe_float(data.get('selling_price'))
        ld = str(data.get('last_digits','')); sd = f"'{ld}" if ld else ""
        # B(2) through M(13) = 12 values
        ws.update(f'B{cell.row}:M{cell.row}', [[
            data.get('card_type',''), sd, data.get('account',''),
            data.get('order_name',''), data.get('order_id',''),
            data.get('model',''), data.get('variant',''),
            cost, sell, sell - cost,
            data.get('delivery_date',''), data.get('sale_month','')
        ]])
        cache_clear('jiomart_orders')
        return jsonify({
            'success': True, 'id': id,
            'card_type': data.get('card_type',''), 'last_digits': ld,
            'account': data.get('account',''),
            'order_name': data.get('order_name',''), 'order_id': data.get('order_id',''),
            'model': data.get('model',''), 'variant': data.get('variant',''),
            'costing': cost, 'selling_price': sell, 'profit': sell - cost,
            'delivery_date': data.get('delivery_date',''),
            'sale_month': data.get('sale_month','')
        })
    except Exception as e: print("Jiomart Edit Error:", e); return jsonify({'success': False})


# ── Exchange Orders API ──────────────────────────────────────────────────────
# id(1) platform(2) model(3) variant(4) costing(5) exchange_model(6)
# exchange_variant(7) exchange_value(8) service_fee(9) original_costing(10)
# last_digits(11) card_type(12) voucher_amount(13) created_at(14)
#
# Key calculation:
# original_costing = (costing - exchange_value) + service_fee

@app.route('/api/exchange-orders', methods=['GET', 'POST'])
def api_exchange_orders():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('exchange_orders')

    if request.method == 'GET':
        cached = cache_get('exchange_orders')
        if cached: return jsonify(cached)
        try:
            records = ws.get_all_records()
        except Exception as e:
            print(f"Exchange Orders GET error: {e}")
            try:
                all_vals = ws.get_all_values()
                if not all_vals or len(all_vals) < 2: return jsonify([])
                headers = all_vals[0]
                records = [dict(zip(headers, row + [''] * (len(headers) - len(row)))) for row in all_vals[1:]]
            except Exception as e2:
                print(f"Exchange GET fallback error: {e2}"); records = []
        for o in records:
            if str(o.get('last_digits', '')).startswith("'"): o['last_digits'] = str(o['last_digits'])[1:]
        result = list(reversed(records))
        cache_set('exchange_orders', result)
        return jsonify(result)

    try:
        data            = request.json; next_id = get_next_id(ws)
        costing         = safe_float(data.get('costing'))
        exchange_value  = safe_float(data.get('exchange_value'))
        service_fee     = safe_float(data.get('service_fee'))
        # original_costing = (costing - exchange_value) + service_fee
        original_costing = round((costing - exchange_value) + service_fee, 2)
        voucher_amount  = safe_float(data.get('voucher_amount'))
        now             = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ld = str(data.get('last_digits', '')); sd = f"'{ld}" if ld else ""
        ws.append_row([
            next_id,
            data.get('platform', ''),
            data.get('model', ''),
            data.get('variant', ''),
            costing if costing else '',
            data.get('exchange_model', ''),
            data.get('exchange_variant', ''),
            exchange_value if exchange_value else '',
            service_fee if service_fee else '',
            original_costing if (costing or exchange_value or service_fee) else '',
            sd,
            data.get('card_type', ''),
            voucher_amount if voucher_amount else '',
            now
        ])
        cache_clear('exchange_orders')
        return jsonify({
            'success': True, 'id': next_id,
            'platform': data.get('platform', ''),
            'model': data.get('model', ''), 'variant': data.get('variant', ''),
            'costing': costing if costing else '',
            'exchange_model': data.get('exchange_model', ''),
            'exchange_variant': data.get('exchange_variant', ''),
            'exchange_value': exchange_value if exchange_value else '',
            'service_fee': service_fee if service_fee else '',
            'original_costing': original_costing if (costing or exchange_value or service_fee) else '',
            'last_digits': ld, 'card_type': data.get('card_type', ''),
            'voucher_amount': voucher_amount if voucher_amount else '',
            'created_at': now
        })
    except Exception as e:
        print(f"Exchange POST Error: {e}"); return jsonify({'success': False}), 500

@app.route('/api/exchange-orders/bulk-delete', methods=['POST'])
def bulk_del_exchange():
    if not SHEET: return jsonify({'success': False})
    ids = request.json.get('ids', []); ws = SHEET.worksheet('exchange_orders')
    try: records = ws.get_all_records()
    except: records = []
    rows_to_delete = [i+2 for i,r in enumerate(records) if r.get('id') in ids]
    for r_idx in sorted(rows_to_delete, reverse=True): ws.delete_row(r_idx)
    cache_clear('exchange_orders')
    return jsonify({'success': True, 'deleted': len(rows_to_delete)})

@app.route('/api/exchange-orders/export')
def export_exchange():
    if not SHEET: return "No sheet connected", 500
    fmt = request.args.get('format', 'csv')
    try: records = SHEET.worksheet('exchange_orders').get_all_records()
    except: records = []
    for o in records:
        if str(o.get('last_digits','')).startswith("'"): o['last_digits'] = str(o['last_digits'])[1:]
    headers = ['id','platform','model','variant','costing','exchange_model','exchange_variant',
               'exchange_value','service_fee','original_costing','last_digits','card_type',
               'voucher_amount','created_at']
    if fmt == 'csv':
        out = io.StringIO(); w = csv.DictWriter(out, fieldnames=headers, extrasaction='ignore')
        w.writeheader(); w.writerows(records); out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode('utf-8')), mimetype='text/csv',
            as_attachment=True, download_name=f'exchange_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.csv')
    wb = openpyxl.Workbook(); ws_xl = wb.active; ws_xl.title = "Exchange Orders"
    hf = PatternFill("solid", fgColor="1A2D45"); hfont = Font(bold=True, color="2ECC8F")
    ws_xl.append(headers)
    for c in ws_xl[1]: c.fill = hf; c.font = hfont
    for r in records: ws_xl.append([r.get(h,'') for h in headers])
    for col in ws_xl.columns:
        ws_xl.column_dimensions[col[0].column_letter].width = min(max((len(str(c.value or '')) for c in col),default=10)+4,40)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'exchange_orders_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')

@app.route('/api/exchange-orders/<int:id>', methods=['DELETE', 'PUT'])
def modify_exchange(id):
    if request.method == 'DELETE':
        cache_clear('exchange_orders'); return delete_master_table('exchange_orders', id)
    if not SHEET: return jsonify({'success': False})
    data = request.json; ws = SHEET.worksheet('exchange_orders')
    try:
        cell            = ws.find(str(id), in_column=1)
        costing         = safe_float(data.get('costing'))
        exchange_value  = safe_float(data.get('exchange_value'))
        service_fee     = safe_float(data.get('service_fee'))
        original_costing = round((costing - exchange_value) + service_fee, 2)
        voucher_amount  = safe_float(data.get('voucher_amount'))
        ld = str(data.get('last_digits', '')); sd = f"'{ld}" if ld else ""
        # B(2) through M(13) = 12 values
        ws.update(f'B{cell.row}:M{cell.row}', [[
            data.get('platform', ''),
            data.get('model', ''), data.get('variant', ''),
            costing if costing else '',
            data.get('exchange_model', ''), data.get('exchange_variant', ''),
            exchange_value if exchange_value else '',
            service_fee if service_fee else '',
            original_costing if (costing or exchange_value or service_fee) else '',
            sd, data.get('card_type', ''),
            voucher_amount if voucher_amount else ''
        ]])
        cache_clear('exchange_orders')
        return jsonify({
            'success': True, 'id': id,
            'platform': data.get('platform', ''),
            'model': data.get('model', ''), 'variant': data.get('variant', ''),
            'costing': costing if costing else '',
            'exchange_model': data.get('exchange_model', ''),
            'exchange_variant': data.get('exchange_variant', ''),
            'exchange_value': exchange_value if exchange_value else '',
            'service_fee': service_fee if service_fee else '',
            'original_costing': original_costing if (costing or exchange_value or service_fee) else '',
            'last_digits': ld, 'card_type': data.get('card_type', ''),
            'voucher_amount': voucher_amount if voucher_amount else ''
        })
    except Exception as e:
        print("Exchange Edit Error:", e); return jsonify({'success': False})


# ── Voucher Tracker API ──────────────────────────────────────────────────────
# id(1) platform(2) amount(3) discount_pct(4) profit(5) month(6) created_at(7)

@app.route('/api/voucher-tracker', methods=['GET', 'POST'])
def api_voucher_tracker():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('voucher_tracker')
    if request.method == 'GET':
        cached = cache_get('voucher_tracker')
        if cached: return jsonify(cached)
        try:
            records = ws.get_all_records()
        except Exception as e:
            print(f"Voucher Tracker GET error (get_all_records): {e}")
            # Fallback: read raw values and map manually using current schema headers
            try:
                all_vals = ws.get_all_values()
                if not all_vals or len(all_vals) < 2:
                    return jsonify([])
                headers = all_vals[0]
                records = []
                for row in all_vals[1:]:
                    # Pad short rows
                    padded = row + [''] * (len(headers) - len(row))
                    records.append(dict(zip(headers, padded)))
            except Exception as e2:
                print(f"Voucher Tracker GET fallback error: {e2}")
                records = []
        result = list(reversed(records))
        cache_set('voucher_tracker', result)
        return jsonify(result)
    try:
        data         = request.json; next_id = get_next_id(ws)
        amount       = safe_float(data.get('amount'))
        disc_pct     = safe_float(data.get('discount_pct'))
        profit       = round(amount * disc_pct / 100, 2)
        is_redeemed  = 1 if data.get('is_redeemed') else 0
        now          = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append_row([
            next_id, data.get('platform',''), data.get('voucher_code',''),
            data.get('voucher_pin',''), amount, disc_pct, profit,
            data.get('month',''), is_redeemed, now
        ])
        cache_clear('voucher_tracker')
        return jsonify({
            'success':True,'id':next_id,
            'platform':data.get('platform',''),'voucher_code':data.get('voucher_code',''),
            'voucher_pin':data.get('voucher_pin',''),
            'amount':amount,'discount_pct':disc_pct,'profit':profit,
            'month':data.get('month',''),'is_redeemed':is_redeemed,'created_at':now
        })
    except Exception as e:
        print(f"Voucher Tracker POST Error: {e}"); return jsonify({'success': False}), 500

@app.route('/api/voucher-tracker/<int:id>', methods=['DELETE', 'PUT'])
def modify_voucher_tracker(id):
    if request.method == 'DELETE':
        cache_clear('voucher_tracker')
        return delete_master_table('voucher_tracker', id)
    # PUT — toggle is_used or update fields
    if not SHEET: return jsonify({'success': False})
    data = request.json; ws = SHEET.worksheet('voucher_tracker')
    try:
        cell         = ws.find(str(id), in_column=1)
        amount       = safe_float(data.get('amount'))
        disc_pct     = safe_float(data.get('discount_pct'))
        profit       = round(amount * disc_pct / 100, 2)
        is_redeemed  = 1 if data.get('is_redeemed') else 0
        # Cols B-I = platform(2) code(3) pin(4) amount(5) disc(6) profit(7) month(8) is_redeemed(9)
        ws.update(f'B{cell.row}:I{cell.row}', [[
            data.get('platform',''), data.get('voucher_code',''),
            data.get('voucher_pin',''), amount, disc_pct, profit,
            data.get('month',''), is_redeemed
        ]])
        cache_clear('voucher_tracker')
        return jsonify({
            'success':True,'id':id,
            'platform':data.get('platform',''),'voucher_code':data.get('voucher_code',''),
            'voucher_pin':data.get('voucher_pin',''),
            'amount':amount,'discount_pct':disc_pct,'profit':profit,
            'month':data.get('month',''),'is_redeemed':is_redeemed
        })
    except Exception as e:
        print("Voucher Tracker PUT Error:", e); return jsonify({'success': False})


# ── Voucher Commission API ───────────────────────────────────────────────────
# id(1) month(2) commission_amount(3) notes(4) created_at(5)
# Tracks monthly commission paid to Pinku from voucher profit

@app.route('/api/voucher-commission', methods=['GET', 'POST'])
def api_voucher_commission():
    if not SHEET: return jsonify([])
    ws = SHEET.worksheet('voucher_commission')
    if request.method == 'GET':
        cached = cache_get('voucher_commission')
        if cached: return jsonify(cached)
        try:
            records = ws.get_all_records()
        except Exception as e:
            print(f"Voucher Commission GET error: {e}")
            try:
                all_vals = ws.get_all_values()
                if not all_vals or len(all_vals) < 2: return jsonify([])
                headers = all_vals[0]
                records = [dict(zip(headers, row + [''] * (len(headers) - len(row)))) for row in all_vals[1:]]
            except: records = []
        cache_set('voucher_commission', records)
        return jsonify(records)
    try:
        data   = request.json; next_id = get_next_id(ws)
        amount = safe_float(data.get('commission_amount'))
        now    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append_row([next_id, data.get('month',''), amount, data.get('notes',''), now])
        cache_clear('voucher_commission')
        return jsonify({'success':True,'id':next_id,'month':data.get('month',''),
                        'commission_amount':amount,'notes':data.get('notes',''),'created_at':now})
    except Exception as e:
        print(f"Voucher Commission POST Error: {e}"); return jsonify({'success':False}), 500

@app.route('/api/voucher-commission/<int:id>', methods=['DELETE','PUT'])
def modify_voucher_commission(id):
    if not SHEET: return jsonify({'success':False})
    if request.method == 'DELETE':
        cache_clear('voucher_commission')
        return delete_master_table('voucher_commission', id)
    data = request.json; ws = SHEET.worksheet('voucher_commission')
    try:
        cell   = ws.find(str(id), in_column=1)
        amount = safe_float(data.get('commission_amount'))
        ws.update(f'B{cell.row}:D{cell.row}',
                  [[data.get('month',''), amount, data.get('notes','')]])
        cache_clear('voucher_commission')
        return jsonify({'success':True,'id':id,'month':data.get('month',''),
                        'commission_amount':amount,'notes':data.get('notes','')})
    except Exception as e:
        print("Voucher Commission PUT Error:", e); return jsonify({'success':False})


# ── Jiomart Migration API ────────────────────────────────────────────────────
# Move selected main_orders rows into jiomart_orders, then delete from main_orders.
# Field mapping:
#   main: card_type last_digits account order_name model variant costing
#         selling_price profit delivery_date sale_month created_at
#   jiomart: card_type last_digits account order_name order_id(blank) model
#            variant costing selling_price profit delivery_date sale_month created_at

@app.route('/api/main-orders/migrate-to-jiomart', methods=['POST'])
def migrate_to_jiomart():
    if not SHEET: return jsonify({'success': False, 'error': 'Not connected'})
    ids          = request.json.get('ids', [])
    delete_after = request.json.get('delete_after', True)
    if not ids: return jsonify({'success': False, 'error': 'No orders selected'})

    try:
        main_ws   = SHEET.worksheet('main_orders')
        jiomart_ws = SHEET.worksheet('jiomart_orders')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

    try:
        main_records = main_ws.get_all_records()
    except Exception as e:
        return jsonify({'success': False, 'error': f'Could not read main_orders: {e}'})

    # Filter selected rows
    to_migrate = [r for r in main_records if r.get('id') in ids]
    if not to_migrate:
        return jsonify({'success': False, 'error': 'No matching orders found'})

    # Get next jiomart id
    next_id = get_next_id(jiomart_ws)
    migrated = 0
    errors   = []

    for r in to_migrate:
        try:
            # Strip the apostrophe prefix gspread adds to last_digits
            ld = str(r.get('last_digits', ''))
            if ld.startswith("'"): ld = ld[1:]
            safe_ld = f"'{ld}" if ld else ""

            costing = safe_float(r.get('costing'))
            selling = safe_float(r.get('selling_price'))
            profit  = safe_float(r.get('profit'))

            jiomart_ws.append_row([
                next_id,
                r.get('card_type', ''),
                safe_ld,
                r.get('account', ''),        # account from main order
                r.get('order_name', ''),      # order_name maps to jiomart order_name
                '',                           # order_id — blank (not in main orders)
                r.get('model', ''),
                r.get('variant', ''),
                costing,
                selling,
                profit,
                r.get('delivery_date', ''),
                r.get('sale_month', 'Current Sale'),
                r.get('created_at', '')
            ])
            next_id += 1
            migrated += 1
        except Exception as e:
            errors.append(f"Order {r.get('id')}: {e}")

    # Delete from main_orders if requested
    deleted = 0
    if delete_after and migrated > 0:
        try:
            # Re-read records to get fresh row numbers after potential appends
            fresh_records = main_ws.get_all_records()
            rows_to_delete = [
                i + 2 for i, rec in enumerate(fresh_records)
                if rec.get('id') in ids
            ]
            for row_idx in sorted(rows_to_delete, reverse=True):
                main_ws.delete_row(row_idx)
                deleted += 1
        except Exception as e:
            errors.append(f"Delete step: {e}")

    cache_clear('main_orders')
    cache_clear('jiomart_orders')

    return jsonify({
        'success': migrated > 0,
        'migrated': migrated,
        'deleted': deleted,
        'errors': errors
    })


@app.route('/api/dashboard-data')
def api_dashboard_data():
    if not SHEET: return jsonify({'error': 'No sheet connected'})
    try:
        def safe_records(tab):
            """Read a sheet tab safely — always returns a list of dicts, never raises."""
            try:
                rows = SHEET.worksheet(tab).get_all_records()
                return rows if rows else []
            except Exception:
                pass
            try:
                all_vals = SHEET.worksheet(tab).get_all_values()
                if not all_vals or len(all_vals) < 2: return []
                headers = all_vals[0]
                # Strip empty trailing headers
                while headers and not headers[-1]: headers.pop()
                records = []
                for row in all_vals[1:]:
                    padded = list(row) + [''] * max(0, len(headers) - len(row))
                    records.append(dict(zip(headers, padded[:len(headers)])))
                return records
            except Exception as e2:
                print(f"Dashboard safe_records({tab}) fallback failed: {e2}")
                return []

        def sf(v):
            try: return float(v) if v not in (None, '', 'None', 'N/A') else 0.0
            except: return 0.0

        def is_sold(o): return sf(o.get('selling_price', 0)) > 0

        # ── FY calculation ──
        from datetime import date
        from collections import defaultdict
        today  = date.today()
        cur_fy = today.year if today.month >= 4 else today.year - 1
        try:    fy = int(request.args.get('fy', cur_fy))
        except: fy = cur_fy
        fy_start = f"{fy}-04"
        fy_end   = f"{fy+1}-03"

        def in_fy(month_str):
            if not month_str: return False
            m = str(month_str)[:7]
            return fy_start <= m <= fy_end

        # Accept optional fy param e.g. ?fy=2024 means Apr 2024 – Mar 2025
        # Default: current financial year
        from datetime import date
        today   = date.today()
        cur_fy  = today.year if today.month >= 4 else today.year - 1
        try:    fy = int(request.args.get('fy', cur_fy))
        except: fy = cur_fy
        fy_start = f"{fy}-04"       # April of fy year
        fy_end   = f"{fy+1}-03"     # March of fy+1

        def in_fy(month_str):
            if not month_str: return False
            m = str(month_str)[:7]
            return fy_start <= m <= fy_end

        main_orders     = safe_records('main_orders')
        sec_orders      = safe_records('secondary_orders')
        offline_orders  = safe_records('offline_orders')
        jiomart_orders  = safe_records('jiomart_orders')
        exchange_orders = safe_records('exchange_orders')
        voucher_tracker = safe_records('voucher_tracker')

        def sf(v): return safe_float(v)
        def is_sold(o): return sf(o.get('selling_price')) > 0

        # Filter everything to current FY
        # Online orders use sale_month field
        def fy_online(orders):
            # Use sale_month if set, fall back to created_at for legacy data
            result = []
            for o in orders:
                m = o.get('sale_month','') or str(o.get('created_at',''))[:7]
                if in_fy(m): result.append(o)
            return result
        def fy_offline(orders):
            return [o for o in orders if in_fy(o.get('sale_month',''))]
        def fy_voucher(vouchers):
            return [v for v in vouchers if in_fy(v.get('month',''))]
        def fy_exchange(orders):
            # exchange uses created_at
            return [o for o in orders if in_fy(str(o.get('created_at',''))[:7])]

        main_fy   = fy_online(main_orders)
        sec_fy    = fy_online(sec_orders)
        jio_fy    = fy_online(jiomart_orders)
        off_fy    = fy_offline(offline_orders)
        exch_fy   = fy_exchange(exchange_orders)
        vouch_fy  = fy_voucher(voucher_tracker)

        online_all  = main_fy + sec_fy + jio_fy
        online_sold = [o for o in online_all  if is_sold(o)]
        off_sold    = [o for o in off_fy      if is_sold(o)]

        online_revenue = sum(sf(o.get('selling_price')) for o in online_sold)
        online_profit  = sum(sf(o.get('profit'))        for o in online_sold)
        online_costing = sum(sf(o.get('costing'))       for o in online_sold)
        off_revenue    = sum(sf(o.get('selling_price')) for o in off_sold)
        off_profit     = sum(sf(o.get('profit'))        for o in off_sold)
        off_costing    = sum(sf(o.get('costing'))       for o in off_sold)

        redeemed   = [v for v in vouch_fy if str(v.get('is_redeemed','0')) == '1']
        pending_v  = [v for v in vouch_fy if str(v.get('is_redeemed','0')) != '1']
        v_red_profit = sum(sf(v.get('profit')) for v in redeemed)
        v_pend_profit= sum(sf(v.get('profit')) for v in pending_v)
        v_face       = sum(sf(v.get('amount')) for v in vouch_fy)
        # Commission paid to Pinku this FY — non-fatal if tab doesn't exist yet
        try:
            all_commission   = safe_records('voucher_commission')
            fy_commission    = [c for c in all_commission if in_fy(c.get('month',''))]
            total_commission = sum(sf(c.get('commission_amount')) for c in fy_commission)
        except Exception:
            fy_commission    = []
            total_commission = 0.0
        net_voucher_profit = round(v_red_profit - total_commission, 2)
        # Monthly commission map for chart
        comm_by_month = {}
        for c in fy_commission:
            m = str(c.get('month',''))[:7]
            if m: comm_by_month[m] = comm_by_month.get(m, 0) + sf(c.get('commission_amount'))

        grand_revenue = online_revenue + off_revenue
        grand_profit  = online_profit  + off_profit  + net_voucher_profit
        grand_costing = online_costing + off_costing

        # ── FY months list Apr→Mar ──
        from collections import defaultdict, OrderedDict
        fy_months = [f"{fy}-{m:02d}" for m in range(4,13)] + [f"{fy+1}-{m:02d}" for m in range(1,4)]
        monthly = {m: {'online_revenue':0,'online_profit':0,'offline_revenue':0,'offline_profit':0,'voucher_profit':0} for m in fy_months}

        for o in online_sold:
            m = str(o.get('sale_month',''))[:7]
            if m in monthly:
                monthly[m]['online_revenue'] += sf(o.get('selling_price'))
                monthly[m]['online_profit']  += sf(o.get('profit'))

        for o in off_sold:
            m = str(o.get('sale_month',''))[:7]
            if m in monthly:
                monthly[m]['offline_revenue'] += sf(o.get('selling_price'))
                monthly[m]['offline_profit']  += sf(o.get('profit'))

        for v in redeemed:
            m = str(v.get('month',''))[:7]
            if m in monthly:
                monthly[m]['voucher_profit'] += sf(v.get('profit'))

        monthly_data = [{'month': m,
            'online_revenue':  round(monthly[m]['online_revenue'],2),
            'online_profit':   round(monthly[m]['online_profit'],2),
            'offline_revenue': round(monthly[m]['offline_revenue'],2),
            'offline_profit':  round(monthly[m]['offline_profit'],2),
            'voucher_profit':  round(monthly[m]['voucher_profit'],2),
            'commission':      round(comm_by_month.get(m, 0), 2),
            'net_voucher':     round(monthly[m]['voucher_profit'] - comm_by_month.get(m, 0), 2),
            'total_profit':    round(monthly[m]['online_profit']+monthly[m]['offline_profit']+monthly[m]['voucher_profit']-comm_by_month.get(m,0),2),
        } for m in fy_months]

        # ── Platform breakdown ──
        plat = defaultdict(lambda: {'count':0,'revenue':0,'profit':0})
        for o in online_sold:
            p = o.get('platform','Unknown') or 'Unknown'
            plat[p]['count']+=1; plat[p]['revenue']+=sf(o.get('selling_price')); plat[p]['profit']+=sf(o.get('profit'))
        platform_data = sorted([{'platform':k,'count':v['count'],'revenue':round(v['revenue'],2),'profit':round(v['profit'],2)} for k,v in plat.items()],key=lambda x:-x['profit'])[:8]

        # ── Top models ──
        model_map = defaultdict(lambda: {'count':0,'profit':0,'revenue':0})
        for o in online_sold + off_sold:
            mn = o.get('model','') or ''
            if mn: model_map[mn]['count']+=1; model_map[mn]['profit']+=sf(o.get('profit')); model_map[mn]['revenue']+=sf(o.get('selling_price'))
        top_models = sorted([{'model':k,'count':v['count'],'profit':round(v['profit'],2),'revenue':round(v['revenue'],2)} for k,v in model_map.items()],key=lambda x:-x['count'])[:8]

        # ── Sale Month summary (online, grouped by month) ──
        month_map = defaultdict(lambda: {'count':0,'sold':0,'profit':0,'pending':0})
        for o in online_all:
            m = str(o.get('sale_month',''))[:7] or 'Unknown'
            month_map[m]['count']+=1
            if is_sold(o): month_map[m]['sold']+=1; month_map[m]['profit']+=sf(o.get('profit'))
            else: month_map[m]['pending']+=1
        month_data = sorted([{'month':k,'count':v['count'],'sold':v['sold'],'pending':v['pending'],'profit':round(v['profit'],2)} for k,v in month_map.items()],key=lambda x:-x.get('month',''))[:12]

        # ── Offline brand breakdown ──
        brand_map = defaultdict(lambda: {'count':0,'profit':0,'revenue':0})
        for o in off_sold:
            b = o.get('brand','') or 'Unknown'
            brand_map[b]['count']+=1; brand_map[b]['profit']+=sf(o.get('profit')); brand_map[b]['revenue']+=sf(o.get('selling_price'))
        brand_data = sorted([{'brand':k,'count':v['count'],'profit':round(v['profit'],2),'revenue':round(v['revenue'],2)} for k,v in brand_map.items()],key=lambda x:-x['count'])[:6]

        # ── Channel split ──
        channel_data = [
            {'channel':'Main Orders',      'count':len(main_fy),  'sold':len([o for o in main_fy  if is_sold(o)]), 'profit':round(sum(sf(o.get('profit')) for o in main_fy  if is_sold(o)),2)},
            {'channel':'Secondary Orders', 'count':len(sec_fy),   'sold':len([o for o in sec_fy   if is_sold(o)]), 'profit':round(sum(sf(o.get('profit')) for o in sec_fy   if is_sold(o)),2)},
            {'channel':'Jiomart',          'count':len(jio_fy),   'sold':len([o for o in jio_fy   if is_sold(o)]), 'profit':round(sum(sf(o.get('profit')) for o in jio_fy   if is_sold(o)),2)},
            {'channel':'Offline',          'count':len(off_fy),   'sold':len(off_sold),                             'profit':round(off_profit,2)},
            {'channel':'Exchange',         'count':len(exch_fy),  'sold':len(exch_fy),                              'profit':0},
        ]

        # ── Available FYs — also check old sale_batch/created_at as fallback ──
        all_months = set()
        for o in main_orders+sec_orders+jiomart_orders:
            # Try sale_month first, fall back to created_at
            m = str(o.get('sale_month','') or o.get('created_at',''))[:7]
            if len(m)==7 and m[4]=='-': all_months.add(m)
        for o in offline_orders:
            m = str(o.get('sale_month',''))[:7]
            if len(m)==7 and m[4]=='-': all_months.add(m)
        fy_set = set()
        for m in all_months:
            try:
                yr,mo = int(m[:4]),int(m[5:7])
                if 2020 <= yr <= 2035: fy_set.add(yr if mo >= 4 else yr-1)
            except: pass
        if not fy_set: fy_set.add(cur_fy)
        available_fys = sorted(fy_set)

        return jsonify({
            'fy': fy, 'fy_label': f"FY {fy}-{str(fy+1)[2:]}",
            'available_fys': available_fys,
            'summary': {
                'grand_revenue': round(grand_revenue,2), 'grand_profit': round(grand_profit,2),
                'grand_costing': round(grand_costing,2),
                'online_orders': len(online_all),   'online_sold': len(online_sold),
                'online_pending': len(online_all)-len(online_sold),
                'offline_orders': len(off_fy),      'offline_sold': len(off_sold),
                'exchange_count': len(exch_fy),     'exch_total_exch': round(sum(sf(o.get('exchange_value')) for o in exch_fy),2),
                'voucher_count':  len(vouch_fy),    'voucher_redeemed_profit': round(v_red_profit,2),
                'voucher_pending_profit': round(v_pend_profit,2), 'voucher_total_face': round(v_face,2),
                'vouchers_redeemed': len(redeemed), 'vouchers_pending': len(pending_v),
                'total_commission': round(total_commission,2),
                'net_voucher_profit': round(net_voucher_profit,2),
                'commission_count': len(fy_commission),
            },
            'monthly': monthly_data, 'platforms': platform_data,
            'top_models': top_models, 'months': month_data,
            'brands': brand_data, 'channels': channel_data,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        # Return a safe minimal response so the dashboard doesn't crash completely
        return jsonify({
            'error': str(e),
            'fy': cur_fy if 'cur_fy' in dir() else 2025,
            'fy_label': 'Error loading data',
            'available_fys': [],
            'summary': {
                'grand_revenue':0,'grand_profit':0,'grand_costing':0,
                'online_orders':0,'online_sold':0,'online_pending':0,
                'offline_orders':0,'offline_sold':0,'exchange_count':0,
                'exch_total_exch':0,'voucher_count':0,'voucher_redeemed_profit':0,
                'voucher_pending_profit':0,'voucher_total_face':0,
                'vouchers_redeemed':0,'vouchers_pending':0,
                'total_commission':0,'net_voucher_profit':0,'commission_count':0,
            },
            'monthly':[],'platforms':[],'top_models':[],'months':[],'brands':[],'channels':[],
        }), 200


@app.route('/setup')
def setup_page(): return render_template('setup.html')

@app.route('/api/setup', methods=['POST'])
def run_setup():
    if not SHEET: return jsonify({'success':False,'error':'Not connected to Google Sheets.'})
    results=[]; existing_titles=[ws.title for ws in SHEET.worksheets()]
    for tab_name, headers in SHEET_SCHEMA.items():
        try:
            if tab_name in existing_titles:
                ws=SHEET.worksheet(tab_name); existing_headers=ws.row_values(1)
                if existing_headers==headers: results.append({'tab':tab_name,'status':'ok','msg':'Already correct'})
                else:
                    ws.update('A1',[headers]); results.append({'tab':tab_name,'status':'fixed','msg':f'Headers updated ({len(headers)} columns)'})
            else:
                ws=SHEET.add_worksheet(title=tab_name,rows=1000,cols=len(headers)+2)
                ws.append_row(headers); results.append({'tab':tab_name,'status':'created','msg':f'Created with {len(headers)} columns'})
        except Exception as e: results.append({'tab':tab_name,'status':'error','msg':str(e)})
    cache_clear_all()
    all_ok=all(r['status'] in ('ok','created','fixed') for r in results)
    return jsonify({'success':all_ok,'results':results})

@app.route('/manifest.json')
def serve_manifest(): return send_file('static/manifest.json',mimetype='application/manifest+json')
@app.route('/sw.js')
def serve_sw(): return send_file('static/sw.js',mimetype='application/javascript')

if __name__=='__main__':
    app.run(host='0.0.0.0',port=5000,debug=True)
